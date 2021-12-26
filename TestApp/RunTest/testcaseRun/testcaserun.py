from TestApp.RunTest.basefile.base import Base
import openpyxl
from pathlib import Path
from luckylog.luckylog import Logger
# from django.contrib.auth.models import User
from django.shortcuts import redirect,render,HttpResponse,reverse
from selenium import webdriver
from TestApp import models
from django.views.decorators.csrf import csrf_exempt
from selenium.webdriver.chrome.options import Options
from django.http import JsonResponse
import json
import time
from TestApp.RunTest.testcaseRun.CreatRresult import creat_result_file
import os
import shutil    #此包用来移动/复制文件
# casefile_Path = Path(__file__).resolve().parents[3]/'media'/'11.xlsx'
# work_book = openpyxl.load_workbook(casefile_Path)
# # sheet = work_book['Sheet1']
# sheets = work_book.sheetnames
# # print(sheets)

@csrf_exempt
def runTest(request):
    current_user = request.user.id
    request_body = json.loads(request.body)
    file = request_body.get('file')
    file_list = models.case_file.objects.filter(title=file,f_num_id=current_user)
    hasReport = str(current_user)+'_'+file
    existReport = models.personcase.objects.filter(reportname=hasReport)
    try:
        existReport.delete()
    except:
        pass
    print(file_list)
    option = Options()
    option.headless = True
    if file_list:
        for case_file in file_list:
            # print(case_file.file)
            wb = Base(webdriver.Chrome(options=option))
            try:
                casefile_Path = Path(__file__).resolve().parents[3] / 'media' / '{}'.format(case_file.file)
                resultDir = Path(__file__).resolve().parents[1] / 'RunedFiles' / 'RunningResult' / '{}_{}'.format(current_user,file)
                reportDir = Path(__file__).resolve().parents[1] / 'RunedFiles' / 'RunningReport' / '{}_{}'.format(current_user,file)
                work_book = openpyxl.load_workbook(casefile_Path)
                print(resultDir,reportDir)
                # sheet = work_book['Sheet1']
                sheets = work_book.sheetnames
                # print(sheets)
                for sheet_name in sheets:
                    sheet = work_book[sheet_name]
                    testFeature = sheet['A1'].value  # 表格的第一个为测试的系统
                    for value_row in sheet.values:
                        # print('用例')
                        if type(value_row[0]) is int:
                            # print('用例')
                            value_list = value_row[2].split(';')
                            Keys_list = []
                            Values_list = []
                            caseName = value_row[3]
                            testStory = value_row[4]    #表格用例的最后一个作为报告的功能点
                            for valueTodict in value_list:
                                kv_list = valueTodict.split('=',1)
                                # print(kv_list)
                                Keys_list.append(kv_list[0])
                                Values_list.append(kv_list[1])
                            # dict_ = dict(zip(Keys_list,Values_list))  #将两个列表合并成一个字典形式
                            dict_ = dict(zip(Keys_list,Values_list))
                            start_time = int(round(time.time()*1000))
                            startTime = time.time()
                            runStatus = getattr(wb,value_row[1])(**dict_)
                            if runStatus:
                                caseStatus = "passed"
                            else:
                                caseStatus = "failed"
                            end_time = int(round(time.time()*1000))
                            # creat_result_file(testFeature=testFeature,testStory=testStory,resultDir=resultDir,resultName=caseName,caseStatus=caseStatus,startTime=start_time,endTime=end_time)
                            endTime = time.time()
                            runTime = round(((endTime*1000)-(startTime*1000))/1000,2)
                            models.personcase.objects.create(user_id=current_user,reportname=str(current_user)+'_'+file,systemName=testFeature,modelName=testStory,caseName=caseName,status=caseStatus,runTime=runTime,caseFile=file)
                wb.quit()
                # os.system('allure generate {0} -o {1} --clean'.format(resultDir,reportDir))
                return JsonResponse({'stu':1})
                # return redirect(reverse('report',kwargs={'filename':file}))
            except Exception as e:
                Logger.erro('程序运行出错，请检查输入的数据',e)
                wb.quit()
                return JsonResponse({'stu':2,'mes':'程序运行出错，请检查数据'})
    else:
        return JsonResponse({'stu':3,'mes':'您还没有用例，请上传用例'})


@csrf_exempt
def push_report(request):
    current_user = request.user.id
    reBody = json.loads(request.body)
    filename = reBody.get('filename')
    reportfile = str(current_user)+'_'+filename
    db_report = models.personcase.objects.filter(reportname=reportfile)
    #定义一个成功用例返回表
    passList = []
    #定义一个失败用例表
    failList = []
    #定义一个汇总的表
    allList = []
    print(reportfile)
    if db_report:
        for case in db_report:
            mysystemName = case.systemName
            modelName = case.modelName
            caseName = case.caseName
            caseStatus = case.status
            runTime = case.runTime
            # list_ = [modelName,caseName,caseStatus,runTime]
            # allList.append(list_)
            if caseStatus == 'passed':
                caseStyle = 'color: #91cc75'
                list_ = [modelName, caseName, caseStatus, runTime, caseStyle]
                allList.append(list_)
                passList.append(list_)
            else:
                caseStyle = 'color: #ee6666'
                list_ = [modelName, caseName, caseStatus, runTime, caseStyle]
                allList.append(list_)
                failList.append(list_)
        allNum = len(passList)+len(failList)
        passingRate = round(len(passList)/allNum,2)*100
        return JsonResponse({'stu':1,'passedNum':passList,'failedNum':failList,'allcaseList':allList,'passRate':passingRate,'mysystemName':mysystemName})
    else:
        return JsonResponse({'stu':2,"mes":'没有找到您的报告文件'})

'''定义一个返回报告文件的API'''
def myreportApi(request):
    if request.method == 'GET':
        currentUser = request.user.id
        reportList = models.personcase.objects.filter(user_id=currentUser)
        print(len(reportList))
        if reportList:
            pushList = []
            for report in reportList:
                myreportName = report.reportname
                mycasefile = report.caseFile
                allPush = [myreportName,mycasefile]
                pushList.append(allPush)
                print(pushList)
            pushreportList = [list(t) for t in set(tuple(_) for _ in pushList)]   #多层嵌套列表去重
            print('xxxxxxxx',pushreportList)
            return JsonResponse({'stu':1,'reportList':pushreportList})
        else:
            return JsonResponse({'stu':2,'mes':'您还没有报告文件'})
    # else:
    #     reBody = json.loads(request.body)
    #     reportName = reBody.get('reportName')
