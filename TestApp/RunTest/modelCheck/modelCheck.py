from pathlib import Path
# from TestApp import views
# from django.forms import models
from django.http.response import JsonResponse
import openpyxl
from django.views.decorators.csrf import csrf_exempt,csrf_protect
from django.contrib.auth.decorators import login_required
from django.shortcuts import render,HttpResponse,redirect
import json
from TestApp import models
import os


keysword_list = ['open','click','input','wait','wait_for','select','iframe','window','js']
@csrf_exempt
def case_Check(request):
    try:
        current_user = request.user.id
        casefile_body = json.loads(request.body)
        user_file = casefile_body.get('casefile')
        file_list = models.case_file.objects.filter(f_num_id=current_user,title=user_file)
        # print(current_user,file_list)
        for fl in file_list:
            case_file = fl.file
            # print(case_file)
            casefile_Path = Path(__file__).resolve().parents[3] / 'media' / '{}'.format(case_file)
            # print(casefile_Path)
            work_book = openpyxl.load_workbook(casefile_Path)
            # sheet = work_book['Sheet1']
            sheets = work_book.sheetnames
            # print(sheets)
            all_case = []
            type_list = []
            type_list2 = []
            for sheet_name in sheets:
                sheet = work_book[sheet_name]
                # print(sheet)
                for value_row in sheet.values:
                    # print(value_row[0])
                    if type(value_row[0]) is int:
                        # print('用例')
                        # type_list.append(value_row[1])
                        if value_row[1] in keysword_list:
                            type_list.append(value_row)
                        else:
                            type_list2.append(value_row)
                work_book.close()
            return JsonResponse({'st':0,'pass_':type_list,'fail_':type_list2})
    except:
        return JsonResponse({'st':1,'mes':'程序出错，请检查数据文件'})
@csrf_exempt
def case_set(request):
        try:
            current_user = request.user.id
            casefile_body = json.loads(request.body)
            user_file = casefile_body.get('casefile')
            file_list = models.case_file.objects.filter(f_num_id=current_user, title=user_file)
            rebody = json.loads(request.body)
            case_id = rebody.get('setcaseid')
            case_type = rebody.get('typevalue')
            # print(case_id,case_type)
            for fl in file_list:
                case_file = fl.file
                casefile_Path = Path(__file__).resolve().parents[3] / 'media' / '{}'.format(case_file)
                work_book = openpyxl.load_workbook(casefile_Path)
                # sheet = work_book['Sheet1']
                sheets = work_book.sheetnames
                # print(sheets)
                all_case = []
                type_list = []
                type_list2 = []
                for sheet_name in sheets:
                    sheet = work_book[sheet_name]
                    rows = 3
                    for value_row in sheet.values:
                        if type(value_row[0]) is int:
                            if value_row[0] == int(case_id):
                                val = 'B{}'.format(rows)
                                # print(val)
                                sheet[val].value = case_type
                                work_book.save(casefile_Path)
                                work_book.close()
                            else:
                                rows+=1
                return JsonResponse({'st':1,"mg":'已修改'})
        except:
            return JsonResponse({'st':0,'mg':'修改失败'})



